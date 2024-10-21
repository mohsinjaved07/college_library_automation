from PyQt5.QtCore import *
from PyQt5.QtGui import *
from PyQt5.QtWidgets import *
from PyQt5.uic import loadUiType
from openpyxl import Workbook
import mysql.connector, sys, datetime


ui, _ = loadUiType("library.ui")


class MainApp(QMainWindow, ui):
    def __init__(self):
        QMainWindow.__init__(self)
        self.setupUi(self)
        self.handle_ui_changes()
        self.handle_ui_buttons()
        self.borrow_return()


    def handle_ui_changes(self):
        self.tabWidget.tabBar().setVisible(False)
        self.tab_2.setEnabled(False)
        self.tab_4.setEnabled(False)
        self.tab_6.setEnabled(False)
        self.tab_5.setEnabled(False)
        self.tab_7.setEnabled(False)
        self.groupBox_2.setEnabled(False)
        self.lineEdit_23.setEnabled(False)
        self.reset_admin_book()
        

    def reset_admin_book(self):
        self.lineEdit_16.setEnabled(False)
        self.lineEdit_17.setEnabled(False)
        self.lineEdit_18.setEnabled(False)
        self.lineEdit_15.setEnabled(False)
        self.lineEdit_19.setEnabled(True)
        self.lineEdit_16.setText('')
        self.lineEdit_17.setText('')
        self.lineEdit_18.setText('')
        self.lineEdit_15.setText('')
        self.lineEdit_19.setText('')
        self.pushButton_10.setEnabled(True)
        self.pushButton_11.setEnabled(False)
        self.pushButton_15.setEnabled(False)


    def handle_ui_buttons(self):
        self.pushButton.clicked.connect(self.user_tab)
        self.pushButton_2.clicked.connect(self.book_tab)
        self.pushButton_3.clicked.connect(self.search_tab)
        self.pushButton_4.clicked.connect(self.return_tab)
        self.pushButton_5.clicked.connect(self.report_tab)
        self.pushButton_14.clicked.connect(self.users_tab)
        self.pushButton_6.clicked.connect(self.add_user)
        self.pushButton_7.clicked.connect(self.login_user)
        self.pushButton_13.clicked.connect(self.update_user)
        self.pushButton_8.clicked.connect(self.login_admin)
        self.pushButton_9.clicked.connect(self.add_books)
        self.pushButton_10.clicked.connect(self.read_books)
        self.pushButton_16.clicked.connect(self.reset_admin_book)
        self.pushButton_11.clicked.connect(self.update_book)
        self.pushButton_15.clicked.connect(self.delete_book)
        self.pushButton_19.clicked.connect(self.search_book)
        self.pushButton_20.clicked.connect(self.borrow_tab)
        self.pushButton_17.clicked.connect(self.borrow_book)
        self.pushButton_31.clicked.connect(self.show_specific_borrowed_book)
        self.pushButton_32.clicked.connect(self.reset_specific_book)
        self.pushButton_30.clicked.connect(self.return_specific_book)
        self.pushButton_12.clicked.connect(self.reportdate_excel)
        self.pushButton_18.clicked.connect(self.specific_reportdate_enable)
        self.pushButton_21.clicked.connect(self.specific_reportdate_excel)
        self.pushButton_22.clicked.connect(self.search_user_info)
        self.pushButton_23.clicked.connect(self.update_user_info)
        self.pushButton_24.clicked.connect(self.delete_user_info)
        self.pushButton_25.clicked.connect(self.userreport_excel)


    def user_tab(self):
        self.tabWidget.setCurrentIndex(0)


    def book_tab(self):
        self.tabWidget.setCurrentIndex(1)


    def search_tab(self):
        self.tabWidget.setCurrentIndex(2)

        if self.tab_4.isEnabled():
            myresult = self.show_books()

            if myresult:
                self.tableWidget_3.setRowCount(0)
                for i in range(0, len(myresult)):
                    self.tableWidget_3.insertRow(i)
                    for j in range(0, len(myresult[i])):
                        self.tableWidget_3.setItem(i, j, QTableWidgetItem(myresult[i][j]))


    def borrow_tab(self):
        book_no = self.label_32.text()
        if book_no == '':
            self.statusBar().showMessage("Please search first.")
        else:
            self.statusBar().showMessage("Here is the information.")
            self.tabWidget.setCurrentIndex(3)


    def return_tab(self):
        self.tabWidget.setCurrentIndex(4)


    def report_tab(self):
        self.tabWidget.setCurrentIndex(5)
        self.comboBox_2.setEnabled(False)
        self.comboBox_3.setEnabled(False)
        self.comboBox_4.setEnabled(False)
        self.comboBox_7.setEnabled(False)
        self.comboBox_6.setEnabled(False)
        self.comboBox_5.setEnabled(False)
        self.pushButton_21.setEnabled(False)


        if self.tab_5.isEnabled():
            self.report_info()


    def users_tab(self):
        self.tabWidget.setCurrentIndex(6)


        if self.tab_7.isEnabled():
            self.reset_user_info()
            myresult = self.show_user_info()
            if myresult:
                self.tableWidget_6.setRowCount(0)
                for i in range(0, len(myresult)):
                    self.tableWidget_6.insertRow(i)
                    for j in range(0, len(myresult[i])):
                        self.tableWidget_6.setItem(i, j, QTableWidgetItem(myresult[i][j]))
    

    def add_user(self):
        mydb = mysql.connector.connect(
            host = 'localhost',
            user = 'root',
            password = '',
            db = 'cla'
        )
        mycursor = mydb.cursor()


        registration_no = self.lineEdit.text()
        name = self.lineEdit_2.text()
        father_name = self.lineEdit_3.text()
        email = self.lineEdit_4.text()
        password = self.lineEdit_5.text()


        if registration_no == '' or name == '' or father_name == '' or email == '' or password == '':
            self.statusBar().showMessage("Invalid Input.")
        else:
            mycursor.execute(f"INSERT INTO user(registration_no, name, father_name, email, password)\
            VALUES('{registration_no}', '{name}', '{father_name}', '{email}', '{password}');")
            self.statusBar().showMessage("User registered successfully.")
            mydb.commit()


        mydb.close()
        self.lineEdit.setText('')
        self.lineEdit_2.setText('')
        self.lineEdit_3.setText('')
        self.lineEdit_4.setText('')
        self.lineEdit_5.setText('')


    def login_user(self):
        mydb = mysql.connector.connect(
            host = 'localhost',
            user = 'root',
            password = '',
            db = 'cla'
        )
        mycursor = mydb.cursor()


        registration_no = self.lineEdit_6.text()
        password = self.lineEdit_7.text()


        mycursor.execute(f"SELECT registration_no, password FROM user\
         WHERE registration_no = '{registration_no}';")
        myresult = mycursor.fetchall()


        if myresult:
            for x in myresult:
                if registration_no == x[0] and password == x[1]:
                    self.statusBar().showMessage("User successfully logged in.")
                    self.groupBox_2.setEnabled(True)
                    self.tab_4.setEnabled(True)
                    self.tab_6.setEnabled(True)
                    self.lineEdit_23.setText(x[0])
                    self.lineEdit_25.setText(x[1])
                    self.label_29.setText(x[0])
                    self.label_33.setText(x[0])
                    self.borrow_return()
                    self.specific_fine()
                    myresult = self.show_books()


                    if myresult:
                        self.tableWidget_3.setRowCount(0)
                        for i in range(0, len(myresult)):
                            self.tableWidget_3.insertRow(i)
                            for j in range(0, len(myresult[i])):
                                self.tableWidget_3.setItem(i, j, QTableWidgetItem(myresult[i][j]))
                else:
                    self.statusBar().showMessage("Invalid password.")
        else:
            self.statusBar().showMessage("Invalid input.")


        mydb.close()
        self.lineEdit_6.setText('')
        self.lineEdit_7.setText('')


    def update_user(self):
        mydb = mysql.connector.connect(
            host = 'localhost',
            user = 'root',
            password = '',
            db = 'cla'
        )
        mycursor = mydb.cursor()


        registration_no = self.lineEdit_23.text()
        password = self.lineEdit_25.text()
        

        if password == '':
            self.statusBar().showMessage("Invalid input.")
        else:
            mycursor.execute(f"UPDATE user SET password='{password}' WHERE registration_no = '{registration_no}';")
            self.statusBar().showMessage("Password updated.")            
            mydb.commit()


        mydb.close()
        self.lineEdit_25.setText('')


    def login_admin(self):
        mydb = mysql.connector.connect(
            host = 'localhost',
            user = 'root',
            password = '',
            db = 'cla'
        )
        mycursor = mydb.cursor()


        name = self.lineEdit_8.text()
        password = self.lineEdit_9.text()


        mycursor.execute(f"SELECT name, password FROM admin\
         WHERE name = '{name}';")
        myresult = mycursor.fetchall()


        if myresult:
            for x in myresult:
                if name == x[0] and password == x[1]:
                    self.statusBar().showMessage("Admin successfully logged in.")
                    self.tab_2.setEnabled(True)
                    self.tab_5.setEnabled(True)
                    self.tab_7.setEnabled(True)
                    myresult = self.show_books()


                    if myresult:
                        self.tableWidget.setRowCount(0)
                        for i in range(0, len(myresult)):
                            self.tableWidget.insertRow(i)
                            for j in range(0, len(myresult[i])):
                                self.tableWidget.setItem(i, j, QTableWidgetItem(myresult[i][j]))
                else:
                    self.statusBar().showMessage("Invalid password.")
        else:
            self.statusBar().showMessage("Invalid input.")


        mydb.close()
        self.lineEdit_8.setText('')
        self.lineEdit_9.setText('')


    def show_books(self):
        mydb = mysql.connector.connect(
            host = 'localhost',
            user = 'root',
            password = '',
            db = 'cla'
        )
        mycursor = mydb.cursor()


        mycursor.execute("SELECT * FROM book;")
        myresult = mycursor.fetchall()
        return myresult


    def read_books(self):
        mydb = mysql.connector.connect(
            host = 'localhost',
            user = 'root',
            password = '',
            db = 'cla'
        )
        mycursor = mydb.cursor()
        book_no = self.lineEdit_19.text()

        if book_no == '':
            self.statusBar().showMessage("Invalid input.")
        else:
            mycursor.execute(f"SELECT book_name, book_publisher, book_author, book_edition FROM book WHERE book_no='{book_no}';")
            myresult = mycursor.fetchall()
            
            if myresult:
                self.lineEdit_16.setEnabled(True)
                self.lineEdit_17.setEnabled(True)
                self.lineEdit_18.setEnabled(True)
                self.lineEdit_15.setEnabled(True)
                self.lineEdit_19.setEnabled(False)
                self.pushButton_11.setEnabled(True)
                self.pushButton_15.setEnabled(True)
                self.pushButton_10.setEnabled(False)


                for x in myresult:
                    self.lineEdit_16.setText(x[0])
                    self.lineEdit_17.setText(x[1])
                    self.lineEdit_18.setText(x[2])
                    self.lineEdit_15.setText(x[3])


                self.statusBar().showMessage("Here is the result.")
            else:
                self.statusBar().showMessage("No book found.")


    def update_book(self):
        mydb = mysql.connector.connect(
            host = 'localhost',
            user = 'root',
            password = '',
            db = 'cla'
        )
        mycursor = mydb.cursor()


        book_no = self.lineEdit_19.text()
        book_name = self.lineEdit_16.text()
        book_publisher = self.lineEdit_17.text()
        book_author = self.lineEdit_18.text()
        book_edition = self.lineEdit_15.text()


        if book_name == '' or book_publisher == '' or book_author == '' or book_edition == '':
            self.statusBar().showMessage("Invalid input.")
        else:
            mycursor.execute(f"UPDATE book SET book_name='{book_name}', book_publisher='{book_publisher}',\
             book_author='{book_author}', book_edition='{book_edition}' WHERE book_no='{book_no}';")
            self.statusBar().showMessage("Book successfully updated.")
            self.reset_admin_book()
            mydb.commit()


        mydb.close()
        myresult = self.show_books()


        if myresult:
            self.tableWidget.setRowCount(0)
            for i in range(0, len(myresult)):
                self.tableWidget.insertRow(i)
                for j in range(0, len(myresult[i])):
                    self.tableWidget.setItem(i, j, QTableWidgetItem(myresult[i][j]))
        else:
            self.tableWidget.clearContents()
            self.tableWidget.removeRow(0)


    def delete_book(self):
        mydb = mysql.connector.connect(
            host = 'localhost',
            user = 'root',
            password = '',
            db = 'cla'
        )
        mycursor = mydb.cursor()


        book_no = self.lineEdit_19.text()


        mycursor.execute(f"DELETE FROM report WHERE book_no = '{book_no}';")
        mycursor.execute(f"DELETE FROM book WHERE book_no='{book_no}';")
        self.statusBar().showMessage("Book successfully deleted.")
        self.reset_admin_book()
        mydb.commit()


        mydb.close()
        myresult = self.show_books()


        if myresult:
            self.tableWidget.setRowCount(0)
            for i in range(0, len(myresult)):
                self.tableWidget.insertRow(i)
                for j in range(0, len(myresult[i])):
                    self.tableWidget.setItem(i, j, QTableWidgetItem(myresult[i][j]))
        else:
            self.tableWidget.clearContents()
            self.tableWidget.removeRow(0)


    def search_book(self):
        mydb = mysql.connector.connect(
            host = 'localhost',
            user = 'root',
            password = '',
            db = 'cla'
        )
        mycursor = mydb.cursor()


        book_name = self.lineEdit_28.text()
        author_name = self.lineEdit_29.text()


        if book_name == '' or author_name == '':
            self.statusBar().showMessage("Invalid input.")
        else:
            mycursor.execute(f"SELECT * FROM book WHERE (book_name, book_author)=('{book_name}', '{author_name}');")
            myresult = mycursor.fetchone()


            if myresult:
                self.listWidget.clear()
                self.tableWidget_3.setRowCount(0)
                self.tableWidget_3.insertRow(0)
                for i in range(0, len(myresult)):
                    self.tableWidget_3.setItem(0, i, QTableWidgetItem(myresult[i]))
                    if i == 0:
                        self.label_32.setText(myresult[i])
                    else:
                        self.listWidget.addItem(myresult[i])

                
                self.statusBar().showMessage("Here is the result.")
            else:
                self.statusBar().showMessage("No book found.")
            

    def add_books(self):
        mydb = mysql.connector.connect(
            host = 'localhost',
            user = 'root',
            password = '',
            db = 'cla'
        )
        mycursor = mydb.cursor()


        book_no = self.lineEdit_11.text()
        book_name = self.lineEdit_14.text()
        book_publisher = self.lineEdit_12.text()
        book_author = self.lineEdit_10.text()
        book_edition = self.lineEdit_13.text()


        if book_no == '' or book_name == '' or book_publisher == '' or book_author == '' or book_edition == '':
            self.statusBar().showMessage("Invalid Input.")
        else:
            mycursor.execute(f"INSERT INTO book(book_no, book_name, book_publisher, book_author, book_edition)\
            VALUES('{book_no}', '{book_name}', '{book_publisher}', '{book_author}', '{book_edition}');")
            self.statusBar().showMessage("Book added successfully.")
            mydb.commit()


        mydb.close()
        self.lineEdit_11.setText('')
        self.lineEdit_14.setText('')
        self.lineEdit_12.setText('')
        self.lineEdit_10.setText('')
        self.lineEdit_13.setText('')
        myresult = self.show_books()


        if myresult:
            self.tableWidget.setRowCount(0)
            for i in range(0, len(myresult)):
                self.tableWidget.insertRow(i)
                for j in range(0, len(myresult[i])):
                    self.tableWidget.setItem(i, j, QTableWidgetItem(myresult[i][j]))


    def borrow_return(self):
        mydb = mysql.connector.connect(
            host = 'localhost',
            user = 'root',
            password = '',
            db = 'cla'
        )
        mycursor = mydb.cursor()


        registration_no = self.label_29.text()
        mycursor.execute(f"SELECT book.book_name, book.book_author, user.name, borrowed_date, due_date, returned_date, fine\
         FROM report INNER JOIN book ON report.book_no=book.book_no INNER JOIN user ON user.registration_no=report.registration_no\
         WHERE report.returned_date IS NULL AND report.registration_no='{registration_no}';")
        myresult = mycursor.fetchall()


        if myresult:
            self.tableWidget_5.setRowCount(0)
            self.tableWidget_2.setRowCount(0)
            for i in range(0, len(myresult)):
                self.tableWidget_5.insertRow(i)
                self.tableWidget_2.insertRow(i)
                for j in range(0, len(myresult[i])):
                    self.tableWidget_5.setItem(i, j, QTableWidgetItem(myresult[i][j]))
                    self.tableWidget_2.setItem(i, j, QTableWidgetItem(myresult[i][j]))
        else:
            self.tableWidget_5.clearContents()
            self.tableWidget_5.removeRow(0)
            self.tableWidget_2.clearContents()
            self.tableWidget_2.removeRow(0)


    def show_specific_borrowed_book(self):
        mydb = mysql.connector.connect(
            host = 'localhost',
            user = 'root',
            password = '',
            db = 'cla'
        )
        mycursor = mydb.cursor()


        registration_no = self.label_33.text()
        book_name = self.lineEdit_57.text()
        author_name = self.lineEdit_58.text()


        if book_name == '' or author_name == '':
            self.statusBar().showMessage("Invalid input.")
        else:
            mycursor.execute(f"SELECT book.book_no, book.book_name, book.book_author, user.name, borrowed_date, due_date, returned_date, fine\
             FROM report INNER JOIN book ON report.book_no=book.book_no INNER JOIN user ON user.registration_no=report.registration_no\
             WHERE report.returned_date IS NULL AND (report.registration_no, book.book_name, book.book_author)=('{registration_no}', '{book_name}', '{author_name}')\
             ;")
            myresult = mycursor.fetchone()


            if myresult:
                self.tableWidget_5.setRowCount(0)
                self.tableWidget_5.insertRow(0)
                for i in range(0, len(myresult)):
                    if i == 0:
                        self.label_44.setText(myresult[i])
                    else:
                        self.tableWidget_5.setItem(0, i - 1, QTableWidgetItem(myresult[i]))

                
                self.statusBar().showMessage("Here is the result.")
            else:
                self.statusBar().showMessage("No book found.")


    def reset_specific_book(self):
        self.lineEdit_57.setText("")
        self.lineEdit_58.setText("")
        self.label_44.setText("")
        self.borrow_return()


    def borrow_book(self):
        mydb = mysql.connector.connect(
            host = 'localhost',
            user = 'root',
            password = '',
            db = 'cla'
        )
        mycursor = mydb.cursor()


        registration_no = self.label_29.text()
        book_no = self.label_32.text()
        day = self.comboBox.currentIndex() + 1
        fromDate = datetime.date.today()
        toDate = fromDate + datetime.timedelta(days=int(day))
        fromDate = str(fromDate)
        toDate = str(toDate)


        mycursor.execute(f"SELECT borrowed_date FROM report WHERE returned_date IS NOT NULL AND (book_no, registration_no, borrowed_date)\
        =('{book_no}', '{registration_no}', '{fromDate}');")
        myresult = mycursor.fetchall()


        if myresult:
            self.statusBar().showMessage("You cannot borrow again on this day. Maybe try next day.")
        else:
            mycursor.execute(f"SELECT book_no, registration_no FROM report WHERE returned_date IS NULL AND (book_no, registration_no)=\
            ('{book_no}', '{registration_no}');")
            myresult = mycursor.fetchall()


            if myresult:
                self.statusBar().showMessage("You have already borrowed this book.")
            else:
                mycursor.execute(f"INSERT INTO report(book_no, registration_no, borrowed_date, due_date)\
                VALUES('{book_no}', '{registration_no}', '{fromDate}', '{toDate}');")
                self.statusBar().showMessage("Book borrowed successfully.")
                mydb.commit()


        mydb.close()
        self.comboBox.setCurrentIndex(0)
        self.borrow_return()


    def return_specific_book(self):
        mydb = mysql.connector.connect(
            host = 'localhost',
            user = 'root',
            password = '',
            db = 'cla'
        )
        mycursor = mydb.cursor()


        registration_no = self.label_33.text()
        book_no = self.label_44.text()
        returned_date = datetime.date.today()
        returned_date = str(returned_date)


        if book_no == '':
            self.statusBar().showMessage("Please search first.")
        else:
            mycursor.execute(f"SELECT * FROM report WHERE returned_date IS NULL AND (book_no, registration_no)=('{book_no}', '{registration_no}')\
             ;")
            myresult = mycursor.fetchall()


            if myresult:
                mycursor.execute(f"UPDATE report SET returned_date='{returned_date}' WHERE (book_no, registration_no)=('{book_no}', '{registration_no}');")
                self.statusBar().showMessage("Book has been successfully returned.")
                mydb.commit()
            else:
                self.statusBar().showMessage("Book has already been returned.")


        mycursor.execute(f"SELECT due_date, returned_date FROM report WHERE (book_no, registration_no)=('{book_no}', '{registration_no}');")
        myresult = mycursor.fetchall()


        if myresult:
            for x in myresult:
                dueDate = x[0]
                dueDate = dueDate.replace("-", "")
                dueDate = datetime.datetime.strptime(dueDate, "%Y%m%d").date()
                returnedDate = x[1]
                returnedDate = returnedDate.replace("-", "")
                returnedDate = datetime.datetime.strptime(returnedDate, "%Y%m%d").date()


                if returnedDate > dueDate:
                    day = 1
                    fine = 50
                    while dueDate != returnedDate: 
                        dueDate = dueDate + datetime.timedelta(days = int(day))
                        fine = fine + 50


                    mycursor.execute(f"UPDATE report SET fine='{fine}' WHERE (book_no, registration_no)=('{book_no}', '{registration_no}') AND fine IS NULL;")
                    mydb.commit()
        else:
            pass
                    
                

        mydb.close()
        self.lineEdit_57.setText('')
        self.lineEdit_58.setText('')
        self.label_44.setText('')
        self.borrow_return()


    def report_info(self):
        mydb = mysql.connector.connect(
            host = 'localhost',
            user = 'root',
            password = '',
            db = 'cla'
        )
        mycursor = mydb.cursor()


        mycursor.execute(f"SELECT book.book_name, book.book_author, user.name, borrowed_date, due_date, returned_date, fine\
         FROM report INNER JOIN book ON report.book_no=book.book_no INNER JOIN user ON user.registration_no=report.registration_no;")
        myresult = mycursor.fetchall()


        if myresult:
            self.tableWidget_4.setRowCount(0)
            for i in range(0, len(myresult)):
                self.tableWidget_4.insertRow(i)
                for j in range(0, len(myresult[i])):
                    self.tableWidget_4.setItem(i, j, QTableWidgetItem(myresult[i][j]))


    def specific_fine(self):
        mydb = mysql.connector.connect(
            host = 'localhost',
            user = 'root',
            password = '',
            db = 'cla'
        )
        mycursor = mydb.cursor()


        registration_no = self.label_29.text()
        mycursor.execute(f"SELECT fine FROM report WHERE fine IS NOT NULL AND registration_no='{registration_no}';")
        myresult = mycursor.fetchall()


        if myresult:
            fine = 0            
            for x in myresult:
                fine = fine + int(x[0])


            fine = str(fine)
            mycursor.execute(f"UPDATE user SET total_fine='{fine}' WHERE registration_no = '{registration_no}';")
            self.label_26.setText(fine)
            mydb.commit()
        else:
            mycursor.execute(f"SELECT total_fine FROM user WHERE registration_no='{registration_no}';")
            myresult = mycursor.fetchone()

            
            if myresult:
                self.label_26.setText(myresult[0])
            else:
                self.label_26.setText('')
                

        mydb.close()


    def reportdate_excel(self):
        mydb = mysql.connector.connect(
            host = 'localhost',
            user = 'root',
            password = '',
            db = 'cla'
        )
        mycursor = mydb.cursor()
        

        mycursor.execute(f"SELECT book.book_name, book.book_author, user.name, borrowed_date, due_date, returned_date, fine\
         FROM report INNER JOIN book ON report.book_no=book.book_no INNER JOIN user ON user.registration_no=report.registration_no;")
        myresult = mycursor.fetchall()


        if myresult:
            title = ('book_name', 'book_author', 'occupied_by', 'borrowed_date', 'due_date', 'returned_date', 'fine')
            myresult.insert(0, title)
            wb = Workbook()
            sheet = wb.active
            sheet.title = 'Book Report'
            sheet.cell(column=1, row=1).value="Book Report List"
            for row in range(1, len(myresult)+1):
                for col in range(1, len(myresult[row-1])+1):
                    sheet.cell(column=col, row=row+1).value = myresult[row-1][col-1]
            wb.save('book_report.xlsx')
            self.statusBar().showMessage("Report file has been created.")
        else:
            self.statusBar().showMessage("No report found.")


        mydb.close()


    def specific_reportdate_enable(self):
        self.comboBox_2.setEnabled(True)
        self.comboBox_3.setEnabled(True)
        self.comboBox_4.setEnabled(True)
        self.comboBox_7.setEnabled(True)
        self.comboBox_6.setEnabled(True)
        self.comboBox_5.setEnabled(True)
        self.pushButton_21.setEnabled(True)


    def specific_reportdate_excel(self):
        mydb = mysql.connector.connect(
            host = 'localhost',
            user = 'root',
            password = '',
            db = 'cla'
        )
        mycursor = mydb.cursor()


        fromDay = self.comboBox_2.currentText()
        fromMonth = self.comboBox_3.currentText()
        fromYear = self.comboBox_4.currentText()
        toDay = self.comboBox_7.currentText()
        toMonth = self.comboBox_6.currentText()
        toYear = self.comboBox_5.currentText()


        mycursor.execute(f"SELECT book.book_name, book.book_author, user.name, borrowed_date, due_date, returned_date, fine\
         FROM report INNER JOIN book ON report.book_no=book.book_no INNER JOIN user ON user.registration_no=report.registration_no\
         WHERE borrowed_date BETWEEN '{fromYear}-{fromMonth}-{fromDay}' AND '{toYear}-{toMonth}-{toDay}';")
        myresult = mycursor.fetchall()


        if myresult:
            title = ('book_name', 'book_author', 'occupied_by', 'borrowed_date', 'due_date', 'returned_date', 'fine')
            myresult.insert(0, title)
            wb = Workbook()
            sheet = wb.active
            sheet.title = 'Book Report'
            sheet.cell(column=1, row=1).value="Book Report List"
            for row in range(1, len(myresult)+1):
                for col in range(1, len(myresult[row-1])+1):
                    sheet.cell(column=col, row=row+1).value = myresult[row-1][col-1]
            wb.save('book_report.xlsx')
            self.statusBar().showMessage("Report file has been created.")
        else:
            self.statusBar().showMessage("No report found.")


        mydb.close()
        self.comboBox_2.setCurrentIndex(0)    
        self.comboBox_3.setCurrentIndex(0)
        self.comboBox_4.setCurrentIndex(0)
        self.comboBox_7.setCurrentIndex(0)
        self.comboBox_6.setCurrentIndex(0)
        self.comboBox_5.setCurrentIndex(0)


    def show_user_info(self):
        mydb = mysql.connector.connect(
            host = 'localhost',
            user = 'root',
            password = '',
            db = 'cla'
        )
        mycursor = mydb.cursor()


        mycursor.execute(f"SELECT * FROM user;")
        myresult = mycursor.fetchall()


        return myresult


    def reset_user_info(self):
        self.lineEdit_22.setEnabled(False)
        self.lineEdit_20.setEnabled(True)
        self.lineEdit_21.setEnabled(False)
        self.lineEdit_26.setEnabled(False)
        self.lineEdit_24.setEnabled(False)
        self.pushButton_23.setEnabled(False)
        self.pushButton_24.setEnabled(False)
        self.pushButton_22.setEnabled(True)
        self.lineEdit_22.setText('')
        self.lineEdit_21.setText('')
        self.lineEdit_26.setText('')
        self.lineEdit_20.setText('')
        self.lineEdit_24.setText('')


    def search_user_info(self):
        mydb = mysql.connector.connect(
            host = 'localhost',
            user = 'root',
            password = '',
            db = 'cla'
        )
        mycursor = mydb.cursor()


        registration_no = self.lineEdit_20.text()


        if registration_no == '':
            self.statusBar().showMessage("Please insert data.")
        else:
            mycursor.execute(f"SELECT name, father_name, email, total_fine FROM user WHERE registration_no='{registration_no}';")
            myresult = mycursor.fetchall()

            if myresult:
                self.lineEdit_22.setEnabled(True)
                self.lineEdit_21.setEnabled(True)
                self.lineEdit_26.setEnabled(True)
                self.lineEdit_20.setEnabled(False)
                self.lineEdit_24.setEnabled(True)
                self.pushButton_23.setEnabled(True)
                self.pushButton_24.setEnabled(True)
                self.pushButton_22.setEnabled(False)
                

                for x in myresult:
                    self.lineEdit_22.setText(x[0])
                    self.lineEdit_21.setText(x[1])
                    self.lineEdit_26.setText(x[2])
                    self.lineEdit_24.setText(x[3])

                self.statusBar().showMessage("Here is the result.")
            else:
                self.statusBar().showMessage("No user found.")


        mydb.close()


    def update_user_info(self):
        mydb = mysql.connector.connect(
            host = 'localhost',
            user = 'root',
            password = '',
            db = 'cla'
        )
        mycursor = mydb.cursor()


        registration_no = self.lineEdit_20.text()
        name = self.lineEdit_22.text()
        father_name = self.lineEdit_21.text()
        email = self.lineEdit_26.text()
        total_fine = self.lineEdit_24.text()


        if name == '' or father_name == '' or email == '':
            self.statusBar().showMessage("Invalid Input.")
        elif total_fine == '':
            mycursor.execute(f"UPDATE user SET name='{name}', father_name='{father_name}', email='{email}' WHERE registration_no='{registration_no}';")
            mydb.commit()
            self.statusBar().showMessage("User updated.")
        else:
            mycursor.execute(f"UPDATE user SET name='{name}', father_name='{father_name}', email='{email}', total_fine='{total_fine}'\
             WHERE registration_no='{registration_no}';")
            self.label_26.setText(total_fine)
            mydb.commit()
            self.statusBar().showMessage("User updated.")


        mydb.close()
        self.reset_user_info()


        myresult = self.show_user_info()
        if myresult:
            self.tableWidget_6.setRowCount(0)
            for i in range(0, len(myresult)):
                self.tableWidget_6.insertRow(i)
                for j in range(0, len(myresult[i])):
                    self.tableWidget_6.setItem(i, j, QTableWidgetItem(myresult[i][j]))
        else:
            self.tableWidget_6.clearContents()
            self.tableWidget_6.removeRow(0)


    def delete_user_info(self):
        mydb = mysql.connector.connect(
            host = 'localhost',
            user = 'root',
            password = '',
            db = 'cla'
        )
        mycursor = mydb.cursor()


        registration_no = self.lineEdit_20.text()
        mycursor.execute(f"DELETE FROM report WHERE registration_no = '{registration_no}';")
        mycursor.execute(f"DELETE FROM user WHERE registration_no = '{registration_no}';")
        self.statusBar().showMessage("User Deleted.")
        mydb.commit()


        self.reset_user_info()
        myresult = self.show_user_info()
        if myresult:
            self.tableWidget_6.setRowCount(0)
            for i in range(0, len(myresult)):
                self.tableWidget_6.insertRow(i)
                for j in range(0, len(myresult[i])):
                    self.tableWidget_6.setItem(i, j, QTableWidgetItem(myresult[i][j]))
        else:
            self.tableWidget_6.clearContents()
            self.tableWidget_6.removeRow(0)


    def userreport_excel(self):
        mydb = mysql.connector.connect(
            host = 'localhost',
            user = 'root',
            password = '',
            db = 'cla'
        )
        mycursor = mydb.cursor()
        

        mycursor.execute(f"SELECT * FROM user;")
        myresult = mycursor.fetchall()


        if myresult:
            title = ('registration_no', 'name', 'father_name', 'email', 'password', 'total_fine')
            myresult.insert(0, title)
            wb = Workbook()
            sheet = wb.active
            sheet.title = 'User Report'
            sheet.cell(column=1, row=1).value="User Report List"
            for row in range(1, len(myresult)+1):
                for col in range(1, len(myresult[row-1])+1):
                    sheet.cell(column=col, row=row+1).value = myresult[row-1][col-1]
            wb.save('user_report.xlsx')
            self.statusBar().showMessage("Report file has been created.")
        else:
            self.statusBar().showMessage("No report found.")


        mydb.close()


def main():
    app = QApplication(sys.argv)
    window = MainApp()
    window.show()
    app.exec()


if __name__ == "__main__":
    main()